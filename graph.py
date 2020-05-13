import openpyxl
from matplotlib import pyplot as plt
from huepy import *

excel_file = str(input(bold(lightgreen('Enter the name of file : '))))

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

number_rows = ws.max_row
number_columns = ws.max_column
column = {} 
i = 1

for col in ws.iter_cols():
    temp = []
    
    for cell in col:
        temp.append(cell.value)

    column[i] = temp
    i +=1

x_axis = columns[1][1:]

for i in range(2,number_columns + 1):
    legend = columns[i][0]
    plt.plot(x_axis, columns[i][1:], label = legend)
    
plt.legend()
plt.show()